import math
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Color

from Objects.Customer import Customer
from Objects.WeeklySession import WeeklySession
from Modules.StringMethod import *
from Modules.jsonStaticFiles import *



        
def formatExcel(path = "LifePlus.xlsx", saveFile = False) -> WeeklySession:
    # Define constants for cell column names
    settings = getSettings()
    NAME        = settings['NAME']
    COMMENT     = settings['COMMENT']
    PRODUCT     = settings['PRODUCT']
    SIZE        = settings['SIZE']
    CAT         = settings['CAT']
    QUANTITY    = settings['QUANTITY']
    PRICE       = settings['PRICE']
    STATUS      = settings['STATUS']
    PHONE       = settings['PHONE']

    wb = load_workbook(path)
    ws = wb['订单列表(行排不合并)']

    #   master object for this excel
    session = WeeklySession()

    wsValues = ws.values

    #   Populate session
    note = ''
    for row in wsValues:
        #   Skip header
        if row[NAME] == '微信昵称':
            continue
        #   Process orders with help from Yonghong
        if row[NAME] == 'Yonghong' and row[COMMENT] != '' and row[COMMENT] is not None: 
            real_name = row[COMMENT]
        else:
            real_name = row[NAME]
            #   prevent note from updating to blank
            if note is None or note.strip() == '':
                if row[COMMENT] is None or str(row[COMMENT]).strip() == '':
                    note = ''
                else:
                    note = row[COMMENT]


        #   Eliminate special characters, except names with only special characters
        real_name = properWords(real_name) if properWords(real_name) != '' else real_name
        real_name = real_name.lower()

        #   process current row
        session.next(real_name, note, row[PRODUCT],row[PRICE],row[QUANTITY])

        #   reset note
        note = ''

    #   Write result to a new sheet
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = '订单列表'

    #   Set Optimized Column width, need to adjust based on actual situation
    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 60
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 10

    ws2.print_options.horizontalCentered = True

    #   Apply Borders and TextWrap
    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    top_left = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'))

    bottom_left = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        bottom=Side(style='thin'))

    #   ----------------------------------Sorting & Ranking--------------------------------
    orderedList: dict [float, Customer] = {}
    #   map a unique number to a customer for ranking
    for customer in session.customerDict.values():
        hashCode: float = 0.0
        #   generate a unique hashcode that represent
        for char in customer.name:
            hashCode += ord(char)
        #   make the magnitude hashCode less significant to avoid anomalies
        #   Basically converting the code to a decimal number
        hashCode /= 10 ** (math.ceil(math.log(hashCode, 10)) + 1)

        #   Load the list with value plus the hash code
        orderedList[round(hashCode + customer.shopList.total,3)] = customer

    #   Now we have the sorted key as a list
    rank = list(orderedList)
    rank.sort(reverse=True)

    #   Load the sorted list into a ordered dictionary in the session
    orderedCustomers: dict[str, Customer] = {}
    for i in rank:
        orderedCustomers[orderedList[i].name] = orderedList[i]

    session.customerDict = orderedCustomers

    #   ----------------------------------Sorting & Ranking ENDS--------------------------------
    
    ROW_PER_PAGE = 50
    rowcounter = 1
    
    def checkRow(row):
        return ROW_PER_PAGE - row % ROW_PER_PAGE

    for cx in session.customerDict.values():
        # - 1 for spaceing, + 1 for comment/header section
        if checkRow(rowcounter) - 1  < len(cx.shopList.shoppingCart) + 1 :
            rowcounter += checkRow(rowcounter) + 1 # + 1 for comment/header
        
        #   Apply styles
        ws2.cell(rowcounter, 1).border = top_left
        ws2.cell(rowcounter, 1).font = Font(bold=True)
        ws2.cell(rowcounter, 2).font = Font(color="00FF0000")
        ws2.cell(rowcounter + 1, 1).alignment = Alignment(horizontal='left')
        ws2.cell(rowcounter, 2).border = Border(top = Side(style='thin'))
        ws2.cell(rowcounter, 3).border = thin_border
        ws2.cell(rowcounter, 4).border = thin_border
        ws2.cell(rowcounter,3).alignment = Alignment(horizontal='center')
        ws2.cell(rowcounter,4).alignment = Alignment(horizontal='center')
        

        ws2.cell(rowcounter, 1).value = cx.name  #   Print customer name
        ws2.cell(rowcounter, 2).value = cx.memo  #   Print customer note
        ws2.cell(rowcounter, 3).value = "单价"                 #    Header
        ws2.cell(rowcounter, 4).value = "数量"                  #    Header
        ws2.cell(rowcounter + 1, 1).value = "总计: ${}".format(cx.shopList.total)  #   Print total price

    
        rowcounter += 1
        
        #   Add additional rows for manual
        # for k in range(1,4):
        #     ws2.cell(rowcounter, 1).border = Border(left = Side(style='thin'))
        #     for i in range(2,5):
        #         ws2.cell(rowcounter, i).border = thin_border
        #     rowcounter += 1

        #   Print each product in shopping list
        for k, v in cx.shopList.shoppingCart.items():
            # Apply borders & Styles
            for i in range(2,5):
                ws2.cell(rowcounter, i).border = thin_border

            ws2.cell(rowcounter, 1).border = Border(left = Side(style='thin'))
            ws2.cell(rowcounter,3).alignment = Alignment(horizontal='center')
            ws2.cell(rowcounter,4).alignment = Alignment(horizontal='center')

            ws2.cell(rowcounter,2).value = k.name
            ws2.cell(rowcounter,3).value = "${}".format(k.price)
            ws2.cell(rowcounter,4).value = v
            rowcounter += 1

        ws2.cell(rowcounter-1, 1).border = bottom_left

        #   Add spaces between each order, if it is not the end of page.
        if checkRow(rowcounter) > 2:
            rowcounter += 1
        
    #   Save file
    newpath = 'LifePlus打印用表' + str(date.today()) + '.xlsx'
    if saveFile:
        wb2.save(newpath)  # Add date to file name

    return session


def addDeliverySheet(result, jsonPath: str = "") -> None:
    rank = result[0][1]
    customer = result[0][0]
    path = result[2]
    wb = load_workbook(path)
    ws = wb.create_sheet("送货清单")

    ws.cell(1, 1).value = '微信昵称'
    ws.cell(1, 2).value = '方位'
    ws.cell(1, 3).value = '总价'
    ws.cell(1, 4).value = 'EMT'
    ws.cell(1, 5).value = '现金'
    ws.cell(1, 6).value = '备注'
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 60

    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
    for i in range(1,7):
        ws.cell(1, i).border = thin_border

    def checkPrice(x):
        return True if x > 120 else False
    
    delivery = filter(checkPrice,rank)

    rowcounter = 2
    for v in delivery:
        
        ws.cell(rowcounter,1).value = customer[v].name
        ws.cell(rowcounter,2).value = ""
        ws.cell(rowcounter,3).value = customer[v].shopList.total
        ws.cell(rowcounter,4).value = ""
        ws.cell(rowcounter,5).value = ""
        ws.cell(rowcounter,6).value = customer[v].memo

        for i in range(1,7):
             ws.cell(rowcounter, i).border = thin_border
        
        ws.cell(rowcounter, 1).font = Font(bold=True)
        ws.cell(rowcounter, 6).font = Font(color="00FF0000")
        ws.cell(rowcounter, 3).alignment = Alignment(horizontal='left')

        rowcounter += 1
    
    wb.save(path)



#   ---------------------------Similarity check-------------------------------
from difflib import SequenceMatcher

def check(a,b):
    return SequenceMatcher(None,a,b).ratio()

def similarityCheck(result):
    customers = result[0][0]

    for n in customers.values():
        for p in customers.values():
            sim = check(n.name, p.name)
            if sim < 1 and sim > 0.6:
                print("Potential overlap: {} <==> {} -----> {}\n".format(n.name, p.name,sim))