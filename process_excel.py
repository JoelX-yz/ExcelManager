'''

'''

import math
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Color

from Objects.customer import Customer
from Objects.group_order import GroupOrder
from Modules.StringMethod import *
from Modules.jsonStaticFiles import *


def format_excel(path: str = "LifePlus.xlsx", space: int = 0, save_file: bool = False) -> GroupOrder:
    """Reads an excel file and combine multiple orders into one

    Args:
        path (str, optional): Path of the excel file. Defaults to "LifePlus.xlsx".
        space (int, optional): Adds additional empty row between each customer's order.  Defaults to 0.
        saveFile (bool, optional): Set the option to generate output file, when performing statistical analysis this should use False. Defaults to False.

    Returns:
        Weeklygroup_order: an object that holds a weekly group order details
    """

    # Define constants for cell column names
    settings = getSettings()
    NAME        = settings['NAME']
    COMMENT     = settings['COMMENT']
    PRODUCT     = settings['PRODUCT']
    SIZE        = settings['SIZE']
    CAT         = settings['CAT']
    QUANTITY    = settings['QUANTITY']
    PRICE       = settings['PRICE']
    STATUS      = settings['STATUS']
    PHONE       = settings['PHONE']

    COL_A_WIDTH = 16
    COL_B_WIDTH = 60
    COL_C_WIDTH = 10
    COL_D_WIDTH = 10

    ROW_PER_PAGE = 50

    wb = load_workbook(path)
    ws = wb['订单列表(行排不合并)']

    #   master object for this excel
    group_order = GroupOrder()

    ws_values = ws.values

    #   Populate group_order
    note = ''
    for row in ws_values:
        #   Skip header
        if row[NAME] == '微信昵称':
            continue
        #   Process orders with help from Yonghong
        if row[NAME] == 'Yonghong' and row[COMMENT] != '' and row[COMMENT] is not None: 
            real_name = row[COMMENT]
        else:
            real_name = row[NAME]

            #   format comment
            if row[COMMENT] is None or str(row[COMMENT]).strip() == '':
                note = ''
            else:
                note = row[COMMENT]


        #   Eliminate special characters, except names with only special characters
        real_name = properWords(real_name) if properWords(real_name) != '' else real_name
        real_name = real_name.lower()

        #   process current row
        group_order.next(real_name, note, row[PRODUCT],row[PRICE],row[QUANTITY])

        #   reset note
        note = ''

    #   Write result to a new sheet
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = '订单列表'

    #   Set Optimized Column width, need to adjust based on actual situation
    ws2.column_dimensions['A'].width = COL_A_WIDTH
    ws2.column_dimensions['B'].width = COL_B_WIDTH
    ws2.column_dimensions['C'].width = COL_C_WIDTH
    ws2.column_dimensions['D'].width = COL_D_WIDTH

    ws2.print_options.horizontalCentered = True

    #   Apply Borders and TextWrap
    thin_border = Border(left   = Side(style='thin'),
                         right  = Side(style='thin'),
                         top    = Side(style='thin'),
                         bottom = Side(style='thin'))

    top_left    = Border(left   = Side(style='thin'),
                         right  = Side(style='thin'),
                         top    = Side(style='thin'))

    bottom_left = Border(left   = Side(style='thin'),
                         right  = Side(style='thin'),
                         bottom = Side(style='thin'))

    #   Sort the dictionary based on the order amount
    group_order.customerDict = dict(sorted(group_order.customerDict.items(),
                                       key=lambda x: x[1].cart.total,
                                       reverse=True))
    
    def calc_empty_row_on_sheet(row):
        return ROW_PER_PAGE - row % ROW_PER_PAGE

    rowcounter = 1

    for customer in group_order.customerDict.values():
        # - 1 for spaceing, + 1 for comment/header section
        if calc_empty_row_on_sheet(rowcounter) - 1  < len(customer.cart.order_items) + 1 :
            rowcounter += calc_empty_row_on_sheet(rowcounter) + 1 # + 1 for comment/header
        
        #   Apply styles
        ws2.cell(rowcounter, 1).border = top_left
        ws2.cell(rowcounter, 1).font = Font(bold=True)
        ws2.cell(rowcounter, 2).font = Font(color="00FF0000")
        ws2.cell(rowcounter + 1, 1).alignment = Alignment(horizontal='left')
        ws2.cell(rowcounter, 2).border = Border(top = Side(style='thin'))
        ws2.cell(rowcounter, 3).border = thin_border
        ws2.cell(rowcounter, 4).border = thin_border
        ws2.cell(rowcounter,3).alignment = Alignment(horizontal='center')
        ws2.cell(rowcounter,4).alignment = Alignment(horizontal='center')
        

        ws2.cell(rowcounter, 1).value = customer.name  
        ws2.cell(rowcounter, 2).value = customer.memo  
        ws2.cell(rowcounter, 3).value = "单价"    
        ws2.cell(rowcounter, 4).value = "数量"    
        ws2.cell(rowcounter + 1, 1).value = "总计: ${}".format(customer.cart.total)  

    
        rowcounter += 1
        
        #  Add additional rows for manual input
        if space > 0:
            for k in range(1, space + 1):
                ws2.cell(rowcounter, 1).border = Border(left = Side(style='thin'))
                for i in range(2,5):
                    ws2.cell(rowcounter, i).border = thin_border
                rowcounter += 1

        #   Print each product in shopping list
        for k, v in customer.cart.order_items.items():
            # Apply borders & Styles
            for i in range(2,5):
                ws2.cell(rowcounter, i).border = thin_border

            ws2.cell(rowcounter,1).border = Border(left = Side(style='thin'))
            ws2.cell(rowcounter,3).alignment = Alignment(horizontal='center')
            ws2.cell(rowcounter,4).alignment = Alignment(horizontal='center')

            ws2.cell(rowcounter,2).value = k.name
            ws2.cell(rowcounter,3).value = "${}".format(k.price)
            ws2.cell(rowcounter,4).value = v
            rowcounter += 1

        ws2.cell(rowcounter-1, 1).border = bottom_left

        #   Add spaces between each order, if it is not the end of page.
        if calc_empty_row_on_sheet(rowcounter) > 2:
            rowcounter += 1
        
    #   Save file
    newpath = 'LifePlus打印用表' + str(date.today()) + '.xlsx'
    group_order.filepath = newpath
    if save_file:
        wb2.save(newpath)  # Add date to file name

    return group_order


def gen_delivery_sheet(current_group_order: GroupOrder) -> None:
    wb = load_workbook(current_group_order.filepath)
    ws = wb.create_sheet("送货清单")

    ws.cell(1, 1).value = '微信昵称'
    ws.cell(1, 2).value = '方位'
    ws.cell(1, 3).value = '总价'
    ws.cell(1, 4).value = 'EMT'
    ws.cell(1, 5).value = '现金'
    ws.cell(1, 6).value = '备注'
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 60

    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
    for i in range(1,7):
        ws.cell(1, i).border = thin_border

    
    delivery = current_group_order.get_delivery_customer()

    rowcounter = 2
    for v in delivery:
        ws.cell(rowcounter,1).value = v.name
        ws.cell(rowcounter,2).value = ""
        ws.cell(rowcounter,3).value = v.cart.total
        ws.cell(rowcounter,4).value = ""
        ws.cell(rowcounter,5).value = ""
        ws.cell(rowcounter,6).value = v.memo

        for i in range(1,7):
             ws.cell(rowcounter, i).border = thin_border
        
        ws.cell(rowcounter, 1).font = Font(bold=True)
        ws.cell(rowcounter, 6).font = Font(color="00FF0000")
        ws.cell(rowcounter, 3).alignment = Alignment(horizontal='left')

        rowcounter += 1
    
    wb.save(current_group_order.filepath)



#   ---------------------------Similarity check-------------------------------
from difflib import SequenceMatcher

def check(a,b):
    return SequenceMatcher(None,a,b).ratio()

def similarityCheck(result):
    customers = result[0][0]

    for n in customers.values():
        for p in customers.values():
            sim = check(n.name, p.name)
            if sim < 1 and sim > 0.6:
                print("Potential overlap: {} <==> {} -----> {}\n".format(n.name, p.name,sim))

def printNames(result):
    customers = result[0][0]

    for n in customers.values():
        print(n.name)